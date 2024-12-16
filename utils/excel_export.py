from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from models import Repair, RepairPart
from sqlalchemy import func

def create_repair_report(repairs, start_date=None, end_date=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Reparaciones"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Encabezados
    headers = ['Código ROV', 'Control', 'Centro', 'Piloto', 'Tipo', 'Falla', 
              'Observaciones', 'Fecha Inicio', 'Fecha Fin', 'Estado', 'Técnico',
              'Repuestos Utilizados', 'Costo Total']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Datos
    for row, repair in enumerate(repairs, 2):
        parts_used = ", ".join([f"{p.part.name} (x{p.quantity})" 
                              for p in repair.parts])
        total_cost = sum(p.quantity * p.unit_cost_at_time for p in repair.parts)

        row_data = [
            repair.rov.code,
            repair.rov.control_code,
            repair.rov.center,
            repair.pilot_name,
            repair.repair_type,
            repair.reported_failure,
            repair.observations,
            repair.start_date.strftime('%d/%m/%Y %H:%M'),
            repair.end_date.strftime('%d/%m/%Y %H:%M') if repair.end_date else 'En proceso',
            repair.status,
            repair.technician.username,
            parts_used,
            f"${total_cost:.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)

    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    return wb

def create_failure_frequency_report(start_date, end_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Frecuencia de Fallas"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Encabezados
    headers = ['Componente', 'Cantidad de Fallas', 'Costo Total', 'Tiempo Promedio de Reparación']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Obtener datos agrupados por componente
    repairs = Repair.query.filter(
        Repair.start_date >= start_date,
        Repair.start_date <= end_date
    ).all()

    component_stats = {}
    for repair in repairs:
        for part in repair.parts:
            component = part.part.name
            if component not in component_stats:
                component_stats[component] = {
                    'failures': 0,
                    'total_cost': 0,
                    'repair_times': []
                }
            
            component_stats[component]['failures'] += 1
            component_stats[component]['total_cost'] += part.quantity * part.unit_cost_at_time
            
            if repair.end_date:
                repair_time = (repair.end_date - repair.start_date).total_seconds() / 3600
                component_stats[component]['repair_times'].append(repair_time)

    # Llenar datos
    row = 2
    for component, stats in component_stats.items():
        avg_time = sum(stats['repair_times']) / len(stats['repair_times']) if stats['repair_times'] else 0
        
        ws.cell(row=row, column=1, value=component)
        ws.cell(row=row, column=2, value=stats['failures'])
        ws.cell(row=row, column=3, value=f"${stats['total_cost']:.2f}")
        ws.cell(row=row, column=4, value=f"{avg_time:.1f} horas")
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        
        row += 1

    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    return wb

def create_monthly_cost_report(year, month):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Costos {month}/{year}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Encabezados
    headers = ['Centro', 'Cantidad Reparaciones', 'Costo Total Repuestos', 
              'Reparaciones Preventivas', 'Reparaciones Correctivas']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Obtener datos por centro
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    repairs = Repair.query.join(Repair.rov).filter(
        Repair.start_date >= start_date,
        Repair.start_date < end_date
    ).all()

    center_stats = {}
    for repair in repairs:
        center = repair.rov.center
        if center not in center_stats:
            center_stats[center] = {
                'total_repairs': 0,
                'total_cost': 0,
                'preventive': 0,
                'corrective': 0
            }
        
        center_stats[center]['total_repairs'] += 1
        if repair.repair_type == 'preventiva':
            center_stats[center]['preventive'] += 1
        else:
            center_stats[center]['corrective'] += 1
        
        for part in repair.parts:
            center_stats[center]['total_cost'] += part.quantity * part.unit_cost_at_time

    # Llenar datos
    row = 2
    for center, stats in center_stats.items():
        ws.cell(row=row, column=1, value=center)
        ws.cell(row=row, column=2, value=stats['total_repairs'])
        ws.cell(row=row, column=3, value=f"${stats['total_cost']:.2f}")
        ws.cell(row=row, column=4, value=stats['preventive'])
        ws.cell(row=row, column=5, value=stats['corrective'])
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        
        row += 1

    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    return wb
